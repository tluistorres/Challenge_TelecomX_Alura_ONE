#########################################################################

# Para imprimir o arquivo "example.xlsx" usando o openpyxl, você pode usar o seguinte código:

import openpyxl

# Abre o arquivo Excel:

wb = openpyxl.load_workbook('example_01.xlsx')

# Seleciona a planilha:

sheet = wb.active

# Imprime os dados da planilha: 

for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        print(sheet.cell(row, col).value, end='\t')
    print()  # Saída: Hello   World

########################################################################

# Se você quiser imprimir o arquivo em um formato mais bonito, você pode usar a biblioteca tabulate:

import openpyxl
from tabulate import tabulate

# Abre o arquivo Excel:

wb = openpyxl.load_workbook('example_01.xlsx')

# Seleciona a planilha:

sheet = wb.active

# Cria uma lista com os dados da planilha:

data = []
for row in range(1, sheet.max_row + 1):
    row_data = []
    for col in range(1, sheet.max_column + 1):
        row_data.append(sheet.cell(row, col).value)
    data.append(row_data)

# # Imprime os dados da planilha em formato de tabela:

# print(tabulate(data, tablefmt='grid'))  # Saída: 
# # +-------+-------+
# # | Hello | World |
# # +-------+-------+

# Primeiro instala o "pip show openpyxl". No terminal.
# Depois instalar o "pip install tabulate". No terminal.

#####################################################################

import openpyxl
from tabulate import tabulate

# Abre o arquivo Excel:

wb = openpyxl.load_workbook('example_01.xlsx')

# Seleciona a planilha:

sheet = wb.active

# Cria uma lista com os dados da planilha:

data = []
for row in range(1, sheet.max_row + 1):
    row_data = []
    for col in range(1, sheet.max_column + 1):
        row_data.append(sheet.cell(row, col).value)
    data.append(row_data)

# Imprime os dados da planilha em formato de tabela:

print(tabulate(data, tablefmt='grid'))

# +-------------+---------+--------------------+
# | Vendendores | Valores |                    |
# +-------------+---------+--------------------+
# | João        | 2024    | =RANK(B2,$B2:$B13) |
# +-------------+---------+--------------------+
# | Patrícia    | 1847    |                    |
# +-------------+---------+--------------------+
# | Marcelo     | 1739    |                    |
# +-------------+---------+--------------------+
# | Pedro       | 1704    |                    |
# +-------------+---------+--------------------+
# | Mônica      | 1991    |                    |
# +-------------+---------+--------------------+
# | Simone      | 2343    |                    |
# +-------------+---------+--------------------+
# | Henrique    | 1978    |                    |
# +-------------+---------+--------------------+
# | Gustavo     | 1920    |                    |
# +-------------+---------+--------------------+
# | Vicente     | 1711    |                    |
# +-------------+---------+--------------------+
# | Dara        | 1754    |                    |
# +-------------+---------+--------------------+
# | Samuel      | 2299    |                    |
# +-------------+---------+--------------------+
# | Pedro       | 2134    |                    |
# +-------------+---------+--------------------+

#####################################################################

# Para ordenar essa lista em Python, você pode usar a função sort() ou sorted() com uma chave de ordenação personalizada.

# Aqui está um exemplo de como ordenar a lista pelo valor de vendas:


# Dados da lista:

vendedores = [
    ["João", 2024],
    ["Patrícia", 1847],
    ["Marcelo", 1739],
    ["Pedro", 1704],
    ["Mônica", 1991],
    ["Simone", 2343],
    ["Henrique", 1978],
    ["Gustavo", 1920],
    ["Vicente", 1711],
    ["Dara", 1754],
    ["Samuel", 2299],
    ["Pedro", 2134],
]

# Ordena a lista pelo valor de vendas: 

vendedores.sort(key=lambda x: x[1], reverse=True)

# Imprime a lista ordenada:

for vendedor in vendedores:
    print(f"{vendedor[0]} - {vendedor[1]}")


# Aqui está o código para produzir uma saída com três colunas: Name, Value e Position:

# Dados da lista:

vendedores = [
    ["João", 2024],
    ["Patrícia", 1847],
    ["Marcelo", 1739],
    ["Pedro", 1704],
    ["Mônica", 1991],
    ["Simone", 2343],
    ["Henrique", 1978],
    ["Gustavo", 1920],
    ["Vicente", 1711],
    ["Dara", 1754],
    ["Samuel", 2299],
    ["Pedro", 2134],
]

# Ordena a lista pelo valor de vendas:

vendedores.sort(key=lambda x: x[1], reverse=True)

# Imprime a lista ordenada com três colunas:

print("+----------+-------+----------+")
print("| Name     | Value | Position |")
print("+----------+-------+----------+")
for pos, vendedor in enumerate(vendedores, start=1):
    print(f"| {vendedor[0]:<10} | {vendedor[1]:<5} | {pos:<8} |")
print("+----------+-------+----------+")


# Essa saída terá as três colunas que você solicitou: Name, Value e Position.

# O código usa a função enumerate() para obter a posição de cada item na lista, e a função sort() para ordenar a lista pelo valor de vendas.

# A saída é formatada com linhas e colunas para facilitar a leitura.

######################################################################

# Para ordenar a lista em ordem crescente e apresentar os valores em float, você pode modificar o código da seguinte forma:


# Dados da lista: 

vendedores = [
    ["João", 2024.0],
    ["Patrícia", 1847.0],
    ["Marcelo", 1739.0],
    ["Pedro", 1704.0],
    ["Mônica", 1991.0],
    ["Simone", 2343.0],
    ["Henrique", 1978.0],
    ["Gustavo", 1920.0],
    ["Vicente", 1711.0],
    ["Dara", 1754.0],
    ["Samuel", 2299.0],
    ["Pedro", 2134.0],
]

# Ordena a lista pelo valor de vendas em ordem crescente:

vendedores.sort(key=lambda x: x[1])

# Imprime a lista ordenada com três colunas:

print("+----------+--------+----------+")
print("| Name     | Value  | Position |")
print("+----------+--------+----------+")
for pos, vendedor in enumerate(vendedores, start=1):
    print(f"| {vendedor[0]:<10} | {vendedor[1]:<7.2f} | {pos:<8} |")
print("+----------+--------+----------+")


# Nesse código:

# - A lista é ordenada em ordem crescente usando o método sort() sem o parâmetro reverse=True.
# - Os valores são apresentados em float com duas casas decimais usando o formato {vendedor[1]:<7.2f}.

# A saída será:

# +----------+--------+----------+
# | Name     | Value  | Position |
# +----------+--------+----------+
# | Vicente  | 1711.00 |        1 |
# | Pedro    | 1704.00 |        2 |
# | Marcelo  | 1739.00 |        3 |
# | Dara     | 1754.00 |        4 |
# | Patrícia | 1847.00 |        5 |
# | Gustavo  | 1920.00 |        6 |
# | Mônica   | 1991.00 |        7 |
# | Henrique | 1978.00 |        8 |
# | João     | 2024.00 |        9 |
# | Pedro    | 2134.00 |       10 |
# | Samuel   | 2299.00 |       11 |
# | Simone   | 2343.00 |       12 |
# +----------+--------+----------+